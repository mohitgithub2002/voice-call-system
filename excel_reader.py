"""
Excel Reader Module
Reads student data from Excel file for fee reminders
"""

import openpyxl
from typing import List, Dict


def read_students(file_path: str) -> List[Dict]:
    """
    Read student data from Excel file.
    
    Expected columns:
    - student_name: Student's name
    - phone_number: Phone with country code (e.g., +919876543210)
    - pending_fees: Amount pending in rupees
    - due_date: Due date for payment
    
    Returns list of student dictionaries.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value.lower().strip() if cell.value else "" for cell in sheet[1]]
        
        # Validate required columns
        required_columns = ['student_name', 'phone_number', 'pending_fees', 'due_date']
        missing = [col for col in required_columns if col not in headers]
        
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        
        # Get column indices
        col_indices = {col: headers.index(col) for col in required_columns}
        
        students = []
        
        # Read data rows (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row[col_indices['student_name']]:
                continue
                
            student = {
                'student_name': str(row[col_indices['student_name']]).strip(),
                'phone_number': str(row[col_indices['phone_number']]).strip(),
                'pending_fees': str(row[col_indices['pending_fees']]).strip(),
                'due_date': str(row[col_indices['due_date']]).strip(),
                'row_number': row_num
            }
            
            # Validate phone number format
            if not student['phone_number'].startswith('+'):
                print(f"⚠️  Row {row_num}: Phone number should start with country code (e.g., +91)")
                # Try to fix Indian numbers
                if student['phone_number'].startswith('91'):
                    student['phone_number'] = '+' + student['phone_number']
                elif len(student['phone_number']) == 10:
                    student['phone_number'] = '+91' + student['phone_number']
            
            students.append(student)
        
        workbook.close()
        return students
        
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")


def create_sample_excel(file_path: str):
    """Create a sample Excel file with correct format."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    
    # Headers
    headers = ['student_name', 'phone_number', 'pending_fees', 'due_date']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    
    # Sample data (you can add your own)
    sample_data = [
        ['राहुल शर्मा', '+919876543210', '5000', '15-02-2026'],
        ['प्रिया गुप्ता', '+919876543211', '7500', '15-02-2026'],
        ['अमित कुमार', '+919876543212', '3000', '20-02-2026'],
    ]
    
    for row_num, data in enumerate(sample_data, start=2):
        for col, value in enumerate(data, start=1):
            sheet.cell(row=row_num, column=col, value=value)
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    
    workbook.save(file_path)
    print(f"✅ Sample Excel created: {file_path}")


if __name__ == "__main__":
    # Test: Create sample file
    create_sample_excel("sample_students.xlsx")
